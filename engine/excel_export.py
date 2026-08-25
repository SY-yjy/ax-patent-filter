"""판정 결과 → 원본 엑셀에 열 추가.

연구원이 이미 쓰고 있는 회차 시트의 **오른쪽 끝에** 결과 열을 붙인다. 새 화면을 익힐
필요 없이 같은 파일에서 필터 걸어 일할 수 있게 하는 것이 목적이다.

  · 회차 시트 오른쪽       — 판정 열 12개 + 과제에 정의된 [수집] 항목 열 (물성 등)
  · 새 시트 `<회차>_AI상세`  — 파고들 때 쓰는 열 (추출 flag/choice·물성·적용규칙·근거인용)

원본 파일은 **절대 수정하지 않는다.** 항상 사본을 만든다(경로 또는 BytesIO).
등급은 서로 배타적이라 등급별 O열 중 하나만 채워지고 나머지는 빈다(요청 형식 유지).
"""
import io
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import criteria as criteria_mod

LABEL_KO = {"noise": "노이즈", "valid": "유효", "issue": "이슈", "hold": "보류"}

# 회차 시트에 붙일 판정 열 (순서 = 왼→오)
MAIN_COLUMNS = ["AI판정", "AI신뢰도",
                "보류", "보류근거", "노이즈", "노이즈근거",
                "유효", "유효근거", "이슈", "이슈근거",
                "근거청구항", "검토필요사항"]


def main_columns(task: dict) -> list[str]:
    """회차 시트에 붙는 실제 열 = 판정 열 + 과제에 정의된 수집 항목 열.

    수집 항목은 과제마다 다르므로 고정 목록으로 둘 수 없다. 상세 시트에만 있어서
    연구원이 같은 시트에서 물성값을 못 보던 문제를 없앤다.
    """
    return MAIN_COLUMNS + [f"[수집] {k}" for k in criteria_mod.collect_names(task)]


def _collected(r: dict) -> dict:
    """추출된 수집값. 예전 실행 기록의 properties/info 도 함께 읽는다."""
    return {**(r.get("properties") or {}), **(r.get("info") or {}), **(r.get("collect") or {})}

_FILL = {"노이즈": "EFEFEF", "유효": "E6F4EA", "이슈": "FCE8E6", "보류": "FEF7E0"}
_HDR_FILL = "D9E2F3"


def _rec_index(results: list[dict]) -> dict:
    """번호 → 판정결과. 번호가 없으면 행 순번(_key idx:N)으로 맞춘다."""
    idx = {}
    for r in results:
        n = r.get("no")
        if n not in (None, ""):
            idx[("no", str(n).split(".")[0])] = r
        key = str(r.get("_key") or "")
        if key.startswith("idx:"):
            idx[("idx", key[4:])] = r
    return idx


def _claims_text(r: dict, limit: int = 900) -> str:
    parts = []
    for c in (r.get("independent_claims") or []):
        no, txt = c.get("no", ""), str(c.get("text", "")).strip()
        parts.append(f"[청구항 {no}] {txt}" if no else txt)
    return "\n\n".join(parts)[:limit]


def _reason_text(r: dict, limit: int = 700) -> str:
    """판정 근거 = 해당한 기준 문장 + 그 근거 인용 + 메모."""
    cid = r.get("criterion_id") or ""
    # 내부 id 는 연구원이 읽는 열에 내보내지 않는다 (추적은 상세 시트의 '기준 id' 열)
    bits = [" ".join(str(r.get("criterion") or r.get("reason") or "").split())]
    q0 = (r.get("evidence_quotes") or {}).get(cid)
    if q0:
        bits.append(f"근거: {str(q0)[:220]}")
    if r.get("notes"):
        bits.append(f"메모: {r['notes']}")
    if not str(r.get("evidence", "")).startswith("원문"):
        bits.append(f"※ 근거등급 {r.get('evidence')} — 원문 미확보")
    return "\n".join(b for b in bits if b)[:limit]


def _main_row(r: dict, task: dict) -> list:
    lab = r.get("label")
    ko = LABEL_KO.get(lab, lab or "")
    reason = _reason_text(r)
    o = {k: "" for k in ("보류", "노이즈", "유효", "이슈")}
    g = {k: "" for k in o}
    if ko in o:
        o[ko] = "O"
        g[ko] = reason
    got = _collected(r)
    return ([ko, r.get("confidence") or "",
             o["보류"], g["보류"], o["노이즈"], g["노이즈"],
             o["유효"], g["유효"], o["이슈"], g["이슈"],
             _claims_text(r), r.get("review_note") or ""]
            + [str(got.get(k, "")) for k in criteria_mod.collect_names(task)])


def _detail_columns(task: dict) -> list[str]:
    cols = criteria_mod.collect_names(task)
    return (["번호", "국가", "발명의 명칭", "AI판정", "AI신뢰도", "기준 id", "해당한 기준",
             "보류사유", "기술요약", "근거등급", "추출경로", "해당O 개수", "모르겠음? 개수"]
            + [f"[수집] {k}" for k in cols]
            + ["근거청구항", "AI 판단과정", "해당한 기준 전체", "모르겠음 기준",
               "근거인용", "추출메모", "검토필요사항", "보정경고"])


def _detail_row(r: dict, task: dict) -> list:
    cols = criteria_mod.collect_names(task)
    got = _collected(r)
    ans = r.get("answers") or {}
    by_id = {c["id"]: c for c in (task.get("criteria") or [])}

    def brief(cid):
        """기준 문장. 지금은 없는 기준(삭제·이름변경)이면 내부 id 대신 그렇게 적는다."""
        c = by_id.get(cid)
        if not c:
            return f"지금은 없는 기준 ({cid})" if not str(cid).startswith("c_") \
                   else "지금은 없는 기준"
        return " ".join(str(c.get("when", cid)).split())[:70]

    hit = [brief(cid) for cid, v in ans.items() if v == "O"]
    unk = [brief(cid) for cid, v in ans.items() if v == "?"]
    q = r.get("evidence_quotes") or {}
    return ([r.get("no"), r.get("country"), r.get("title"),
             LABEL_KO.get(r.get("label"), r.get("label") or ""), r.get("confidence") or "",
             r.get("criterion_id") or "", " ".join(str(r.get("criterion") or "").split())[:200],
             (" ".join(str(r.get("criterion") or "").split())[:80]
              if r.get("label") == "hold" else ""), r.get("tech_summary") or "",
             r.get("evidence") or "", r.get("claims_via") or "",
             r.get("n_matched"), r.get("n_unknown")]
            + [str(got.get(k, "")) for k in cols]
            + [_claims_text(r, 2000),
               criteria_mod.humanize(str(r.get("narration") or ""), task)[:3000],
               "\n".join(hit)[:1500],
               "\n".join(unk)[:1000],
               "\n\n".join(f"{brief(k)}\n  → {v}" for k, v in q.items())[:2000],
               r.get("notes") or "", r.get("review_note") or "",
               "; ".join(r.get("normalize_warnings") or [])[:500]])


def annotate_workbook(src, sheet: str, results: list[dict], task: dict,
                      out_path=None) -> str | bytes:
    """원본 워크북을 열어 `sheet` 오른쪽에 결과 열 + 상세 시트를 붙인다.

    src      : 원본 경로 또는 file-like(BytesIO)
    out_path : 주면 그 경로로 저장하고 경로를 반환, 없으면 xlsx bytes 를 반환(웹 다운로드용)
    """
    wb = openpyxl.load_workbook(src)
    if sheet not in wb.sheetnames:
        raise KeyError(f"시트 없음: {sheet} (있는 시트: {wb.sheetnames})")
    ws = wb[sheet]

    # 헤더행에서 '번호' 열 위치를 찾아 결과를 행에 맞춘다 (위치 가정 금지)
    header = [str(c.value or "").replace("\n", "").replace(" ", "").strip()
              for c in ws[1]]
    no_col = header.index("번호") + 1 if "번호" in header else 1

    start = ws.max_column + 1
    idx = _rec_index(results)

    # 헤더 쓰기
    mcols = main_columns(task)
    for j, name in enumerate(mcols):
        c = ws.cell(row=1, column=start + j, value=name)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=_HDR_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 데이터 쓰기
    written = 0
    for row in range(2, ws.max_row + 1):
        raw_no = ws.cell(row=row, column=no_col).value
        key = ("no", str(raw_no).split(".")[0]) if raw_no not in (None, "") else ("idx", str(row - 1))
        r = idx.get(key) or idx.get(("idx", str(row - 1)))
        if not r:
            continue
        for j, v in enumerate(_main_row(r, task)):
            cell = ws.cell(row=row, column=start + j, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ko = LABEL_KO.get(r.get("label"), "")
        if ko in _FILL:                        # AI판정 열에 등급 색
            ws.cell(row=row, column=start).fill = PatternFill("solid", fgColor=_FILL[ko])
        written += 1

    # 열 너비 · 자동필터 확장 · 틀고정 유지
    widths = {"AI판정": 10, "AI신뢰도": 9, "보류": 6, "노이즈": 7, "유효": 6, "이슈": 6,
              "보류근거": 46, "노이즈근거": 46, "유효근거": 46, "이슈근거": 46,
              "근거청구항": 60, "검토필요사항": 34}
    for j, name in enumerate(mcols):
        ws.column_dimensions[get_column_letter(start + j)].width = (
            widths.get(name, 22 if name.startswith("[수집]") else 14))
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # 상세 시트 (있으면 갈아끼움)
    dname = f"{sheet}_AI상세"
    if dname in wb.sheetnames:
        del wb[dname]
    dws = wb.create_sheet(dname)
    dcols = _detail_columns(task)
    dws.append(dcols)
    for c in dws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=_HDR_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in results:
        dws.append(_detail_row(r, task))
    dws.freeze_panes = "D2"
    dws.auto_filter.ref = f"A1:{get_column_letter(len(dcols))}{dws.max_row}"
    for j, name in enumerate(dcols, start=1):
        w = 40 if name in ("기술요약", "기준 문장", "근거청구항", "AI 판단과정",
                           "근거인용", "추출메모", "검토필요사항", "해당한 기준 전체",
                           "모르겠음 기준") else \
            (24 if name.startswith("[수집]") else 14)
        dws.column_dimensions[get_column_letter(j)].width = w

    if out_path:
        Path(out_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return str(out_path)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
